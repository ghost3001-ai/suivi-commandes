import os
import re
from collections import OrderedDict, defaultdict
from functools import lru_cache

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CATALOG_PATH = os.path.join(BASE_DIR, 'data', 'categorie_ass1.xlsx')
DEFAULT_FAMILY_OVERRIDE_PATH = os.path.join(BASE_DIR, 'data', 'familles_ass1.xlsx')

FAMILY_RE = re.compile(r'^\d+\.\s*([0-9A-Z]+)\b')
PLAIN_FAMILY_RE = re.compile(r'^([0-9]{1,2}X?)\s*[-:]\s+')
CATEGORY_RE = re.compile(r'^(?:Nouveau\s*:\s*)?([0-9]{3,4}[A-Z]?)\s*[-:]\s+')


def _fix_mojibake(value):
    if not value or not any(marker in value for marker in ('Ã', 'Â', 'â')):
        return value
    try:
        return value.encode('latin1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value


def _clean_label(value):
    if value is None:
        return None

    text = _fix_mojibake(str(value))
    text = text.replace('\xa0', ' ').replace('–', '-').replace('—', '-')
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'\s*-\s*', ' - ', text)
    text = re.sub(r'\s*:\s*', ': ', text)
    return text or None


def _extract_family_code(label):
    if not label:
        return None

    match = FAMILY_RE.match(label)
    if match:
        return match.group(1)

    match = PLAIN_FAMILY_RE.match(label)
    if match:
        return match.group(1)
    return None


def _extract_category_code(label):
    if not label:
        return None

    match = CATEGORY_RE.match(label)
    if match:
        return match.group(1)
    return None


def _register_family(families, family_order, raw_label):
    label = _clean_label(raw_label)
    code = _extract_family_code(label)
    if not label or not code:
        return None

    if code not in families:
        families[code] = {
            'label': label,
            'categories': OrderedDict(),
        }
        family_order.append(code)
    return code


def _register_category(
    families,
    category_to_family,
    canonical_category_labels,
    family_code,
    raw_label,
):
    label = _clean_label(raw_label)
    code = _extract_category_code(label)
    if not family_code or not label or not code or family_code not in families:
        return None

    canonical_label = canonical_category_labels.get(code) or label
    families[family_code]['categories'].setdefault(code, canonical_label)
    canonical_category_labels.setdefault(code, canonical_label)
    category_to_family[code] = family_code
    return canonical_label


def _infer_family_code(category_code, families):
    if not category_code:
        return None

    numeric_part = ''.join(ch for ch in category_code if ch.isdigit())
    if not numeric_part:
        return None

    first_three = int(numeric_part[:3]) if len(numeric_part) >= 3 else int(numeric_part)

    range_rules = [
        (100, 200, '10X'),
        (200, 300, '21'),
        (300, 400, '30'),
        (400, 430, '40X'),
        (430, 440, '43X'),
        (500, 511, '50'),
        (511, 520, '51'),
        (600, 700, '60'),
        (700, 720, '70'),
        (720, 800, '7X'),
        (800, 900, '80'),
        (900, 980, '90'),
        (980, 1000, '99'),
    ]
    for lower, upper, family_code in range_rules:
        if lower <= first_three < upper and family_code in families:
            return family_code

    best_match = None
    best_length = -1
    for family_code in families:
        prefix = ''.join(ch for ch in family_code if ch.isdigit())
        if not prefix:
            continue
        if family_code.endswith('X'):
            if numeric_part.startswith(prefix) and len(prefix) > best_length:
                best_match = family_code
                best_length = len(prefix)
        elif numeric_part.startswith(prefix) and len(prefix) > best_length:
            best_match = family_code
            best_length = len(prefix)
    return best_match


def _parse_hierarchy_sheet(
    workbook,
    sheet_name,
    family_col,
    category_col,
    families,
    family_order,
    category_to_family,
    canonical_category_labels,
):
    if sheet_name not in workbook.sheetnames:
        return

    worksheet = workbook[sheet_name]
    current_family_code = None

    for row in worksheet.iter_rows(values_only=True):
        family_value = _clean_label(row[family_col]) if len(row) > family_col else None
        category_value = _clean_label(row[category_col]) if len(row) > category_col else None

        family_code = _extract_family_code(family_value)
        category_code = _extract_category_code(category_value)

        if family_code and not category_code:
            current_family_code = _register_family(families, family_order, family_value)
            continue

        if category_code and current_family_code:
            _register_category(
                families,
                category_to_family,
                canonical_category_labels,
                current_family_code,
                category_value,
            )


def _parse_mapping_sheet(
    workbook,
    sheet_name,
    families,
    category_to_family,
    canonical_category_labels,
    legacy_by_category,
):
    if sheet_name not in workbook.sheetnames:
        return

    worksheet = workbook[sheet_name]
    headers = None

    for row in worksheet.iter_rows(values_only=True):
        values = [_clean_label(value) for value in row]
        if not any(values):
            continue

        if 'Catégorie article' in values and 'Nouvelle Famille' in values:
            headers = values
            continue

        if not headers:
            continue

        row_data = dict(zip(headers, values))
        category_label = _clean_label(row_data.get('Nouvelle Famille'))
        legacy_label = _clean_label(row_data.get('Catégorie article'))
        category_code = _extract_category_code(category_label)

        if not category_code:
            continue

        family_code = category_to_family.get(category_code) or _infer_family_code(category_code, families)
        if family_code:
            canonical_label = _register_category(
                families,
                category_to_family,
                canonical_category_labels,
                family_code,
                category_label,
            )
        else:
            canonical_label = canonical_category_labels.get(category_code) or category_label
            canonical_category_labels.setdefault(category_code, canonical_label)

        if legacy_label and canonical_label:
            legacy_by_category[canonical_label].add(legacy_label)


def _build_empty_catalog(path):
    return {
        'source_path': path,
        'families': [],
        'categories_by_family': {},
        'subcategories_by_category': {},
        'family_lookup': set(),
        'category_lookup': set(),
        'category_to_family': {},
    }


def _load_family_overrides(path=None):
    override_path = path or DEFAULT_FAMILY_OVERRIDE_PATH
    if not override_path or not os.path.exists(override_path):
        return OrderedDict()

    workbook = load_workbook(override_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    families = OrderedDict()

    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            label = _clean_label(value)
            code = _extract_family_code(label)
            if label and code and code not in families:
                families[code] = label

    return families


def _apply_family_overrides(families, category_to_family, family_overrides):
    if not family_overrides:
        return families, list(families.keys()), {}

    overridden_families = OrderedDict(
        (
            family_code,
            {
                'label': family_label,
                'categories': OrderedDict(),
            },
        )
        for family_code, family_label in family_overrides.items()
    )

    category_records = []
    for family_code, family_data in families.items():
        for category_code, category_label in family_data['categories'].items():
            category_records.append(
                (
                    category_code,
                    category_label,
                    category_to_family.get(category_code) or family_code,
                )
            )

    rebuilt_category_to_family = {}
    for category_code, category_label, existing_family_code in category_records:
        target_family_code = existing_family_code
        if target_family_code not in overridden_families:
            target_family_code = _infer_family_code(category_code, overridden_families)
        if not target_family_code or target_family_code not in overridden_families:
            continue

        overridden_families[target_family_code]['categories'].setdefault(category_code, category_label)
        rebuilt_category_to_family[category_label] = overridden_families[target_family_code]['label']

    return overridden_families, list(overridden_families.keys()), rebuilt_category_to_family


def load_category_catalog(path=None, family_override_path=None):
    source_path = path or DEFAULT_CATALOG_PATH
    if not os.path.exists(source_path):
        return _build_empty_catalog(source_path)

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    families = OrderedDict()
    family_order = []
    category_to_family = {}
    canonical_category_labels = {}
    legacy_by_category = defaultdict(set)

    _parse_hierarchy_sheet(
        workbook,
        'Feuil3',
        0,
        1,
        families,
        family_order,
        category_to_family,
        canonical_category_labels,
    )
    _parse_hierarchy_sheet(
        workbook,
        'nouvelle codification ASS',
        1,
        2,
        families,
        family_order,
        category_to_family,
        canonical_category_labels,
    )
    _parse_hierarchy_sheet(
        workbook,
        'categorie afrilux',
        1,
        2,
        families,
        family_order,
        category_to_family,
        canonical_category_labels,
    )

    for sheet_name in ('Produit Afrilux', 'Produit Smart', 'Stock ASS'):
        _parse_mapping_sheet(
            workbook,
            sheet_name,
            families,
            category_to_family,
            canonical_category_labels,
            legacy_by_category,
        )

    family_overrides = _load_family_overrides(family_override_path)
    families, family_order, category_to_family_label = _apply_family_overrides(
        families,
        category_to_family,
        family_overrides,
    )

    family_labels = [families[family_code]['label'] for family_code in family_order]
    categories_by_family = {}

    for family_code in family_order:
        family_label = families[family_code]['label']
        category_labels = list(families[family_code]['categories'].values())
        categories_by_family[family_label] = category_labels
        for category_label in category_labels:
            category_to_family_label.setdefault(category_label, family_label)

    # Les sous-catégories ne sont plus exploitées dans l'application.
    subcategories_by_category = {}

    return {
        'source_path': source_path,
        'families': family_labels,
        'categories_by_family': categories_by_family,
        'subcategories_by_category': subcategories_by_category,
        'family_lookup': set(family_labels),
        'category_lookup': set(category_to_family_label.keys()),
        'category_to_family': category_to_family_label,
    }


@lru_cache(maxsize=2)
def get_category_catalog(path=None, family_override_path=None):
    return load_category_catalog(path, family_override_path)
